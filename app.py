from flask import Flask, request, jsonify, send_file, render_template
import pandas as pd
import io
import os
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024  # 200MB

# Company mapping (Y&S vs Non Y&S)
COMPANY_MAPPING = {
    "Damon and Crew": "Non Y&S",
    "The Ticket Guy": "Non Y&S",
    "YourTickets": "Non Y&S",
    "GK LLC": "Y&S",
    "Jacks YS": "Y&S",
    "Levovitz": "Y&S",
    "Needle Tickets LLC": "Y&S",
    "Pollak Tickets": "Y&S",
    "Yoni Levine": "Y&S",
    "YS Katz": "Y&S",
    "YS Tickets": "Y&S",
    "YS TL": "Y&S",
    "YSA": "Y&S",
    "YSA 2": "Y&S",
    "YSA 3": "Y&S",
    "YSM Tickets": "Y&S",
    "YSS Tickets": "Y&S",
    "YS-Seatgeek": "Y&S",
    "YS-Seatgeek2": "Y&S",
    "YSW": "Y&S",
}

def get_main_company(company):
    """Map company to Main Company name."""
    if company in ("YS-Seatgeek", "YS-Seatgeek2"):
        return "YS Tickets"
    if company in ("YSA 2", "YSA 3"):
        return "YSA"
    return company

def get_ys_mapping(company):
    """Return Y&S or Non Y&S for a company."""
    return COMPANY_MAPPING.get(company, "Non Y&S")

def style_header_row(ws, num_cols):
    """Apply header styling."""
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_data_rows(ws, num_rows, num_cols):
    """Apply alternating row colors and font."""
    light_fill = PatternFill("solid", fgColor="DCE6F1")
    for row in range(2, num_rows + 2):
        fill = light_fill if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

def auto_fit_columns(ws):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 40)

def build_excel_output(df, sheet_name="Available"):
    """Build a styled Excel file from a dataframe."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Write headers
    headers = list(df.columns)
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)

    # Write data
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            # Format cost/total cost columns as currency
            col_name = headers[ci - 1]
            if col_name in ("Cost", "Total Cost"):
                cell.number_format = '#,##0.00'

    style_header_row(ws, len(headers))
    style_data_rows(ws, len(df), len(headers))
    auto_fit_columns(ws)

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    files = request.files.getlist("files")
    month_end_date = request.form.get("month_end_date", "").strip()

    if not files or all(f.filename == "" for f in files):
        return jsonify({"error": "No files uploaded."}), 400
    if not month_end_date:
        return jsonify({"error": "Please enter a month end date."}), 400

    dfs = []
    for f in files:
        if f.filename.endswith(".xlsx"):
            try:
                df = pd.read_excel(f, dtype=str)
                dfs.append(df)
            except Exception as e:
                return jsonify({"error": f"Could not read {f.filename}: {str(e)}"}), 400

    if not dfs:
        return jsonify({"error": "No valid Excel files found."}), 400

    # Combine
    combined = pd.concat(dfs, ignore_index=True)

    # Ensure Cost and Quantity are numeric
    combined["Cost"] = pd.to_numeric(combined["Cost"], errors="coerce")
    combined["Quantity"] = pd.to_numeric(combined["Quantity"], errors="coerce")

    # Add Total Cost column right after Cost
    combined["Total Cost"] = combined["Quantity"] * combined["Cost"]

    # Reorder: insert Total Cost after Cost
    cols = list(combined.columns)
    cost_idx = cols.index("Cost")
    cols.remove("Total Cost")
    cols.insert(cost_idx + 1, "Total Cost")
    combined = combined[cols]

    # Add Main Company column at the beginning
    combined.insert(0, "Main Company", combined["Company"].apply(get_main_company))

    # Add Y&S mapping column (used for splitting, not included in output)
    combined["_ys_flag"] = combined["Company"].apply(get_ys_mapping)

    # Split into Y&S and Non Y&S
    df_ys = combined[combined["_ys_flag"] == "Y&S"].drop(columns=["_ys_flag"])
    df_non_ys = combined[combined["_ys_flag"] == "Non Y&S"].drop(columns=["_ys_flag"])

    # Build filenames
    fname_ys = f"Inventory - {month_end_date} (YS).xlsx"
    fname_non_ys = f"Inventory - {month_end_date} (Non YS).xlsx"

    buf_ys = build_excel_output(df_ys, sheet_name="Available")
    buf_non_ys = build_excel_output(df_non_ys, sheet_name="Available")

    # Package into zip
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(fname_ys, buf_ys.read())
        zf.writestr(fname_non_ys, buf_non_ys.read())
    zip_buf.seek(0)

    response = send_file(
        zip_buf,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"Inventory - {month_end_date}.zip"
    )
    response.headers["X-YS-Rows"] = str(len(df_ys))
    response.headers["X-NonYS-Rows"] = str(len(df_non_ys))
    response.headers["X-Total-Rows"] = str(len(combined))
    response.headers["Access-Control-Expose-Headers"] = "X-YS-Rows, X-NonYS-Rows, X-Total-Rows"
    return response

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
